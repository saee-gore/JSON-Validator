# Validated_Data is the data we are validating against!!
import json
import copy
from openpyxl import load_workbook

def readData(filename, page_number, start_row, end_row, end_col):
    """
    :param file_name: str(input name of the FILE)
    :param page_number: int (sheet number of the FIlE)
    :param start_row: int (starting row number of the sheet)
    :param end_row: int (ending row number of the sheet)
    :param end_col: int (ending col number of the sheet)
    :return: List of the input data is being stored from the input file
    """
    wb = load_workbook(filename)
    origin_file = wb.worksheets[page_number]
    col_info = origin_file[1]
    item = {}
    list_of_items = []
    for x in range(start_row, end_row + 1):
        for y in range(1, end_col + 1):
            item.update(
                {col_info[y - 1].value: origin_file.cell(row=x, column=y).value})
        list_of_items.append(copy.copy(item))
        item.clear()
    return list_of_items


def findLastRowOfExcel(file_name, sheet_name):
    """
    :param file_name: str(input name of the FILE)
    :param sheet_name: int (sheet number of the FIlE)
    :return: int
    """
    wb = load_workbook(file_name)
    lastRowVal = wb[sheet_name].max_row
    return lastRowVal


def main():
    """
    :return: Strings [If any errors, strings are printed in the terminal]
    """
    latestError = ""  # string storing the specific error message
    Errors = """"""  # prints out all the errors in the terminal

    for index in range(0, endRow-1):
        jsoncompanyId = data[index]["companyId"]
        jsoncompanyName = data[index]["companyName"]
        jsoncompanyCategory = data[index]["companyCategory"]

        if jsoncompanyId != ExcelData[index]['companyId']:
            indexErr = str(index+2)
            latestError = "\n Error at "+indexErr+". check data for " +str(jsoncompanyId)+" "+str(ExcelData[index]['companyId'])
            Errors = Errors+latestError+"\n"

        if jsoncompanyName != ExcelData[index]['companyName']:
            indexErr = str(index+2)
            latestError = "\n Error at "+indexErr +  ". check data for "+" " +  str(ExcelData[index]['companyName'])+" and "+jsoncompanyName
            Errors = Errors+latestError+"\n"

        if jsoncompanyCategory != ExcelData[index]['companyCategory']:
            indexErr = str(index+2)
            latestError = "\n Error at "+indexErr +  ". check data for " + " " + str(ExcelData[index]['companyCategory'])+" and "+jsoncompanyCategory
            Errors = Errors+latestError+"\n"
    
    if len(Errors)==0:
        print("Success! All data is good :)")
    else:
        print("All nodes are goood except:"+Errors)


if __name__ == "__main__":
    excelFileData = "BusinessInputData.xlsx"
    endRow = findLastRowOfExcel(excelFileData, "Sheet1")
    ExcelData = readData(excelFileData, 0, 2, endRow, 3)
    jsonFileName = open('input.json')  # Opening JSON file
    data = json.load(jsonFileName)  # returns JSON object as a dictionarindex
    main()
